import openpyxl
import re
import logging


def column_extractor(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[wb.sheetnames[0]]
    status_column = tuple(sheet['J:J'])
    index_column = tuple(sheet['A:A'])
    wb.close()
    return zip(index_column, status_column)


def splitter(column_tuple):
    big_list = {}
    for index_id, status_cell in column_tuple:
        if not index_id.value or not status_cell.value:
            break
        plan_list = re.split(r"(\d{1,2}/\d{1,2} -)", status_cell.value)
        plan_list = [x + y for x, y in zip(plan_list[1::2], plan_list[2::2])]
        big_list[index_id.value] = plan_list
    return big_list


def rewriter(xl_path, new_entry_dic):
    wb1 = openpyxl.load_workbook(xl_path)
    wb2 = openpyxl.load_workbook('/Users/jamlo/Desktop/Blank.xlsx')
    sheet1 = wb1[wb1.sheetnames[0]]
    sheet2 = wb2[wb2.sheetnames[0]]
    q = 2
    for row in sheet1.iter_rows():
        if not row[1].value:
            break
        for new_data in new_entry_dic[row[0].value]:
            new_row = [cell.value for cell in row] + [new_data]
            for col, val in enumerate(new_row, start=1):
                sheet2.cell(row=q, column=col).value = val
            q += 1
    wb1.close()
    wb2.save('/Users/jamlo/Desktop/Blank1.xlsx')


def main():
    x = column_extractor('/Users/jamlo/Desktop/export_summary_1.xlsx')
    logging.debug('Extracted columns from orignal file')
    y = splitter(x)
    logging.debug('Split text into distinct values')
    rewriter('/Users/jamlo/Desktop/export_summary_1.xlsx', y)
    logging.debug('Wrote to file. Done.')


if __name__ == '__main__':
    main()

