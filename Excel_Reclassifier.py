import openpyxl


def column_extractor(file_path):
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[wb.sheetnames[0]]
    pgrad_column = tuple(sheet['I:I'])
    email_column = tuple(sheet['B:B'])
    wb.close()
    return zip(email_column, pgrad_column)


def splitter(column_tuple):
    big_list = {}
    for email, cell in column_tuple:
        if not email.value or not cell.value:
            break
        plan_list = cell.value.split(',')
        big_list[email.value] = plan_list
    return big_list


def rewriter(xl_path, new_entry_dic):
    wb1 = openpyxl.load_workbook(xl_path)
    wb2 = openpyxl.load_workbook('/Users/jamlo/Desktop/Blank.xlsx')
    sheet1 = wb1[wb1.sheetnames[0]]
    sheet2 = wb2[wb2.sheetnames[0]]
    q = 2
    for row in sheet1.iter_rows():
        if not row[1].value:
            break
        for add, new_data in enumerate(new_entry_dic[row[1].value]):
            new_row = [cell.value for cell in row] + [new_data]
            for col, val in enumerate(new_row, start=1):
                sheet2.cell(row=q, column=col).value = val
            q += 1
    wb1.close()
    wb2.save('/Users/jamlo/Desktop/Filled.xlsx')


def main():
    x = column_extractor('/Users/jamlo/Downloads/Career Choice Survey (Responses).xlsx')
    y = splitter(x)
    rewriter('/Users/jamlo/Downloads/Career Choice Survey (Responses).xlsx', y)


if __name__ == '__main__':
    main()
